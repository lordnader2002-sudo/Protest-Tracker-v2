#!/usr/bin/env python3
"""
Protest Tracker
===============
Queries the Mobilize.us public API for events near every property in
properties.csv and produces a professional Excel workbook.

Two reporting windows
---------------------
  • 3-Day Events  : Any public events (protest / rally / town-hall …)
                    within 3 miles of a property, starting today through +3 days.
  • No Kings (30d): Events whose title or description contains a "No Kings"
                    keyword, within 3 miles, today through +30 days.

Usage
-----
  python3 protest_tracker.py [--output FILENAME] [--csv PROPERTIES_CSV]
"""

import argparse
import csv
import json
import math
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

_EASTERN = ZoneInfo("America/New_York")

import requests
try:
    from tqdm import tqdm as _tqdm
    _HAS_TQDM = True
except ImportError:
    _HAS_TQDM = False

try:
    import zipcodes as _zipcodes
    _HAS_ZIPCODES = True
except ImportError:
    _HAS_ZIPCODES = False

_zip_latlon_cache: dict[str, tuple[float, float] | None] = {}

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────

PROPERTIES_CSV  = "properties.csv"
MOBILIZE_API    = "https://api.mobilize.us/v1/events"

# Optional API key — set via env var MOBILIZE_API_KEY or --api-key argument.
# Authenticated requests have their own rate-limit bucket (avoids shared-IP 429s).
MOBILIZE_API_KEY = os.environ.get("MOBILIZE_API_KEY", "")

SEARCH_RADIUS_MI  = 3    # hard filter: must be within this many miles of property
CLUSTER_RADIUS_MI = 35   # properties within this distance share one API query
DAYS_GENERAL      = 3    # 3-day event window
DAYS_NO_KINGS     = 30   # No Kings event window

NO_KINGS_KEYWORDS = ["no kings", "nokings", "no_kings", "#nokings", "50501"]

# Event types to EXCLUDE from all sheets (pure campaign/admin work, meetings).
EXCLUDE_TYPES = {"PHONE_BANK", "TEXT_BANK", "AUTOMATED_PHONE_BANK", "LETTER_WRITING",
                 "VOTER_REG", "FUNDRAISER", "TRAINING", "FRIEND_TO_FRIEND_OUTREACH",
                 "MEETING"}

REQUEST_DELAY  = 2.0   # seconds between successful API requests
MAX_RETRIES    = 3     # max retries on transient errors (NOT 429)
RETRY_BACKOFF  = 2.0   # exponential backoff base (seconds): 2, 4, 8
PER_PAGE       = 200   # max results per page

# ── Excel colour palette ───────────────────────────────────────────────────────

HDR_NAVY   = PatternFill("solid", fgColor="1F3864")   # dark navy  (title rows)
HDR_BLUE   = PatternFill("solid", fgColor="2E75B6")   # medium blue (column headers)
HDR_PURPLE = PatternFill("solid", fgColor="5B2C8D")   # purple      (No Kings title)
COL_PURPLE = PatternFill("solid", fgColor="7B5EA7")   # lighter purple (No Kings col hdrs)
ALT_BLUE   = PatternFill("solid", fgColor="DCE6F1")   # pale blue   (even data rows)
ALT_PURPLE = PatternFill("solid", fgColor="E8DEFF")   # pale purple (even No Kings rows)
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

# Distance-based row shading
FILL_RED   = PatternFill("solid", fgColor="FFCCCC")   # < 1 mile
FILL_AMBER = PatternFill("solid", fgColor="FFE8A0")   # 1–2 miles
FILL_GREEN = PatternFill("solid", fgColor="CCFFCC")   # > 2 miles
FILL_NOTES = PatternFill("solid", fgColor="F0F0F0")   # Notes column

THIN = Border(
    left   = Side(style="thin", color="B8CCE4"),
    right  = Side(style="thin", color="B8CCE4"),
    top    = Side(style="thin", color="B8CCE4"),
    bottom = Side(style="thin", color="B8CCE4"),
)

_THICK_SIDE = Side(style="medium", color="2C3E50")
_THIN_INNER = Side(style="thin",   color="AAAAAA")


def _group_border(top_edge: bool, bot_edge: bool,
                  left_edge: bool, right_edge: bool) -> Border:
    """Thick on group perimeter, light-gray inside."""
    return Border(
        top    = _THICK_SIDE if top_edge   else _THIN_INNER,
        bottom = _THICK_SIDE if bot_edge   else _THIN_INNER,
        left   = _THICK_SIDE if left_edge  else _THIN_INNER,
        right  = _THICK_SIDE if right_edge else _THIN_INNER,
    )

# ── Sheet column definitions ───────────────────────────────────────────────────

COLUMNS = [
    # (header label,            row-dict key,      col width)
    ("Property Name",           "property_name",    28),
    ("Event Title",             "event_title",      42),
    ("Event Type",              "event_type",       16),
    ("Date & Time (EST)",       "event_date",       22),
    ("Event Location",          "event_location",   44),
    ("Distance (mi)",           "distance_mi",      14),
    ("Event URL",               "event_url",        12),
    ("Is New?",                 "is_new",            8),
    ("Duplicate?",              "is_duplicate",      10),
    ("Recurring?",              "is_recurring",      10),
    ("First Seen",              "first_seen",       12),
    ("Notes",                   "notes",            60),
]

# ── Helpers ────────────────────────────────────────────────────────────────────

def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance in miles between two lat/lon points."""
    R = 3958.8
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


def pad_zip(raw: str) -> str:
    """Zero-pad a zipcode to 5 digits (handles New England codes stored as ints)."""
    return str(raw).strip().zfill(5)


def load_properties(path: str) -> list[dict]:
    props = []
    with open(path, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            try:
                props.append({
                    "id":      row["property_id"].strip(),
                    "name":    row["name"].strip(),
                    "address": row["address"].strip(),
                    "zip":     pad_zip(row["postal_code"]),
                    "lat":     float(row["lat"]),
                    "lon":     float(row["lon"]),
                })
            except (ValueError, KeyError):
                pass   # skip malformed rows
    return props


class RateLimitError(Exception):
    """Raised when Mobilize.us returns HTTP 429 so the caller can retry later."""


def _build_headers() -> dict:
    headers = {"User-Agent": "ProtestTracker/1.0 (internal security monitoring tool)"}
    if MOBILIZE_API_KEY:
        headers["Authorization"] = f"Bearer {MOBILIZE_API_KEY}"
    return headers


def fetch_events_for_zip(zipcode: str, max_dist: int,
                         start_ts: int, end_ts: int) -> list[dict]:
    """Return all public Mobilize events for a zip within the given time window.

    On HTTP 429 (rate-limit): skips the zip immediately — no hanging retries.
    On transient errors: retries up to MAX_RETRIES times with exponential backoff.
    """
    params: dict = {
        "zipcode":        zipcode,
        "max_dist":       max_dist,
        "timeslot_start": f"gte_{start_ts}",
        "timeslot_end":   f"lte_{end_ts}",
        "per_page":       PER_PAGE,
        "visibility":     "PUBLIC",
    }
    events: list[dict] = []
    url: str | None = MOBILIZE_API
    headers = _build_headers()

    while url:
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                resp = requests.get(
                    url,
                    params=params if url == MOBILIZE_API else None,
                    headers=headers,
                    timeout=20,
                )

                if resp.status_code == 429:
                    # Signal caller to retry this cluster later.
                    raise RateLimitError(zipcode)

                if resp.status_code == 400:
                    # Bad request — invalid zip or malformed params. Not retryable.
                    print(f"    [400] Bad request for zip={zipcode} — skipping cluster.",
                          file=sys.stderr)
                    return events

                resp.raise_for_status()
                payload = resp.json()
                events.extend(payload.get("data", []))
                url = payload.get("next")
                params = None
                time.sleep(REQUEST_DELAY)
                break   # success — move to next page

            except requests.RequestException as exc:
                wait = RETRY_BACKOFF ** attempt
                print(f"    [WARN] API error zip={zipcode} (attempt {attempt}/{MAX_RETRIES}): "
                      f"{exc}. Retrying in {wait:.0f}s …", file=sys.stderr)
                time.sleep(wait)
        else:
            # All retries exhausted for this page (transient errors only)
            print(f"    [ERROR] Giving up on zip={zipcode} after {MAX_RETRIES} attempts.",
                  file=sys.stderr)
            url = None

    return events


def event_location_str(loc: dict) -> str:
    parts = []
    if loc.get("venue"):
        parts.append(loc["venue"])
    addr = ", ".join(a for a in loc.get("address_lines", []) if a)
    if addr:
        parts.append(addr)
    city_state = ", ".join(filter(None, [loc.get("locality"), loc.get("region")]))
    if city_state:
        parts.append(city_state)
    if loc.get("postal_code"):
        parts.append(loc["postal_code"])
    return ", ".join(parts) if parts else "Virtual / TBD"


def expand_event(event: dict, prop: dict, stats: dict | None = None) -> list[dict]:
    """
    Expand a Mobilize event into one row per timeslot, filtered to
    SEARCH_RADIUS_MI from the property.  Returns [] if outside radius
    or location is unavailable.
    """
    loc = event.get("location") or {}
    # Mobilize nests coordinates under location.location.latitude / .longitude
    nested = loc.get("location") or {}
    elat = nested.get("latitude") or loc.get("lat")
    elon = nested.get("longitude") or loc.get("lon")
    if elat is None or elon is None:
        # Last resort: fall back to zip centroid
        postal = (loc.get("postal_code") or "").strip()[:5]
        if postal and _HAS_ZIPCODES:
            if postal not in _zip_latlon_cache:
                matches = _zipcodes.matching(postal)
                if matches:
                    _zip_latlon_cache[postal] = (float(matches[0]["lat"]),
                                                  float(matches[0]["long"]))
                else:
                    _zip_latlon_cache[postal] = None
            coords = _zip_latlon_cache.get(postal)
            if coords:
                elat, elon = coords
            else:
                if stats is not None:
                    stats["no_coords"] += 1
                return []
        else:
            if stats is not None:
                stats["no_coords"] += 1
            return []   # virtual / no postal code

    dist = haversine(prop["lat"], prop["lon"], elat, elon)
    if dist > SEARCH_RADIUS_MI:
        if stats is not None:
            stats["too_far"] += 1
        return []

    title      = (event.get("title") or "").strip()
    etype      = (event.get("event_type") or "OTHER").replace("_", " ").title()
    browser_url = event.get("browser_url", "")
    loc_str    = event_location_str(loc)

    eid = event.get("id")
    created_unix = event.get("created_date")
    first_seen = (
        datetime.fromtimestamp(created_unix, tz=timezone.utc)
        .astimezone(_EASTERN).strftime("%b %d, %Y")
        if created_unix else ""
    )
    sponsor_name = ((event.get("sponsor") or {}).get("name") or "").strip()
    e_lat = round(float(elat), 6) if elat is not None else ""
    e_lon = round(float(elon), 6) if elon is not None else ""
    rows = []
    for ts in event.get("timeslots") or []:
        start_unix = ts.get("start_date")
        if not start_unix:
            continue
        dt_utc = datetime.fromtimestamp(start_unix, tz=timezone.utc)
        rows.append({
            "event_id":       eid,
            "property_id":    prop["id"],
            "property_name":  prop["name"],
            "property_addr":  prop["address"],
            "event_title":    title,
            "event_type":     etype,
            "event_date":     dt_utc.astimezone(_EASTERN).strftime("%b %d, %Y  %H:%M EST"),
            "event_dt_sort":  dt_utc,
            "event_location": loc_str,
            "distance_mi":    round(dist, 2),
            "event_url":      browser_url,
            "first_seen":     first_seen,
            "sponsor_name":   sponsor_name,
            "event_lat":      e_lat,
            "event_lon":      e_lon,
            "prop_lat":       round(float(prop["lat"]), 6),
            "prop_lon":       round(float(prop["lon"]), 6),
        })
    return rows


def is_no_kings(event: dict) -> bool:
    text = (
        (event.get("title") or "") + " " + (event.get("description") or "")
    ).lower()
    return any(kw in text for kw in NO_KINGS_KEYWORDS)


def get_prev_event_ids(cache_path: str) -> set | None:
    """Return the set of event IDs from a previous cache file, or None if unavailable."""
    if not os.path.exists(cache_path):
        return None
    try:
        with open(cache_path) as f:
            data = json.load(f)
        ids = {r["event_id"] for sheet in ("general", "no_kings")
               for r in data.get(sheet, []) if r.get("event_id")}
        return ids or None
    except Exception:
        return None


def annotate_event_flags(rows: list[dict], prev_event_ids: set | None = None) -> None:
    """Add is_new, is_duplicate, is_recurring fields to each row in-place.

    is_new       – event_id not present in the previous run's cache.
    is_duplicate – same event appears near 2+ distinct properties this run.
    is_recurring – same event has 2+ distinct timeslots this run.
    """
    from collections import defaultdict

    event_props = defaultdict(set)   # event_id → set of property_ids
    event_slots = defaultdict(set)   # event_id → set of timeslots

    for row in rows:
        eid = row.get("event_id")
        if eid:
            event_props[eid].add(row.get("property_id", ""))
            event_slots[eid].add(row.get("event_dt_sort"))

    for row in rows:
        eid = row.get("event_id")
        row["is_new"]       = "Yes" if (prev_event_ids is not None
                                        and eid and eid not in prev_event_ids) else ""
        row["is_duplicate"] = "Yes" if eid and len(event_props[eid]) > 1 else ""
        row["is_recurring"] = "Yes" if eid and len(event_slots[eid]) > 1 else ""


def collapse_recurring_timeslots(rows: list[dict]) -> list[dict]:
    """For each (event_id, property_id) pair keep only the single nearest upcoming
    timeslot.  The is_recurring flag is preserved so the reader still knows the
    event repeats.  Non-recurring rows pass through unchanged.
    """
    from collections import defaultdict

    _FAR_FUTURE = datetime(9999, 12, 31, tzinfo=timezone.utc)

    # Group row indices by (event_id, property_id)
    groups: dict[tuple, list[int]] = defaultdict(list)
    for i, row in enumerate(rows):
        key = (row.get("event_id"), row.get("property_id", ""))
        groups[key].append(i)

    keep: set[int] = set()
    for indices in groups.values():
        if len(indices) == 1:
            keep.add(indices[0])
        else:
            # Nearest = smallest event_dt_sort
            best = min(indices,
                       key=lambda i: rows[i].get("event_dt_sort") or _FAR_FUTURE)
            keep.add(best)

    # Preserve original sort order
    return [row for i, row in enumerate(rows) if i in keep]


# ── Geographic clustering ──────────────────────────────────────────────────────

def cluster_properties(properties: list[dict]) -> list[list[dict]]:
    """
    Group properties into geographic clusters so nearby properties share a
    single API query instead of one query per zip code.

    Algorithm: greedy single-linkage — seed each cluster from the first
    unassigned property and absorb any unassigned property within
    CLUSTER_RADIUS_MI of that seed.  Returns a list of clusters, where
    each cluster is a list of property dicts.
    """
    assigned: set[int] = set()
    clusters: list[list[dict]] = []

    for i, seed in enumerate(properties):
        if i in assigned:
            continue
        cluster = [seed]
        assigned.add(i)
        for j, candidate in enumerate(properties):
            if j in assigned:
                continue
            if haversine(seed["lat"], seed["lon"],
                         candidate["lat"], candidate["lon"]) <= CLUSTER_RADIUS_MI:
                cluster.append(candidate)
                assigned.add(j)
        clusters.append(cluster)

    return clusters


def cluster_query_params(cluster: list[dict]) -> tuple[str, int]:
    """
    Return (zip_code, query_radius_miles) for a cluster.

    zip  – taken from the property nearest to the cluster's geographic centroid
    radius – distance from centroid to farthest member + SEARCH_RADIUS_MI + 1 mile buffer
    """
    centroid_lat = sum(p["lat"] for p in cluster) / len(cluster)
    centroid_lon = sum(p["lon"] for p in cluster) / len(cluster)

    nearest = min(cluster,
                  key=lambda p: haversine(centroid_lat, centroid_lon, p["lat"], p["lon"]))
    max_dist = max(haversine(centroid_lat, centroid_lon, p["lat"], p["lon"])
                   for p in cluster)
    query_radius = math.ceil(max_dist + SEARCH_RADIUS_MI + 1)

    return nearest["zip"], query_radius


# ── Core collection logic ──────────────────────────────────────────────────────

def collect_events(
    properties: list[dict],
) -> tuple[list[dict], list[dict]]:
    """
    Cluster properties geographically, then fire one Mobilize.us API query
    per cluster instead of one per zip code.  Typical reduction: ~200 calls
    down to ~40-60 calls.

    Returns:
      general_rows  – protest-type events within DAYS_GENERAL days
      no_kings_rows – No Kings events within DAYS_NO_KINGS days
    Both lists are sorted by event start time.
    """
    now_ts  = int(datetime.now(tz=timezone.utc).timestamp())
    end_30d = now_ts + DAYS_NO_KINGS * 86_400
    end_3d  = now_ts + DAYS_GENERAL  * 86_400

    clusters = cluster_properties(properties)
    print(f"  Clustered {len(properties)} properties into {len(clusters)} geographic groups.\n")

    general_rows:  list[dict] = []
    no_kings_rows: list[dict] = []
    seen_general:  set[tuple] = set()
    seen_no_kings: set[tuple] = set()

    stats = {"no_coords": 0, "too_far": 0, "excluded_type": 0,
             "outside_window": 0, "passed": 0}

    def process_events(events: list[dict], cluster: list[dict]) -> None:
        for ev in events:
            eid       = ev.get("id")
            etype_raw = (ev.get("event_type") or "OTHER").upper()
            no_kings  = is_no_kings(ev)
            for prop in cluster:
                rows = expand_event(ev, prop, stats)
                if not rows:
                    continue
                for row in rows:
                    ts_key = (eid, prop["id"], row["event_dt_sort"])
                    if etype_raw in EXCLUDE_TYPES:
                        stats["excluded_type"] += 1
                    elif row["event_dt_sort"].timestamp() > end_3d:
                        stats["outside_window"] += 1
                    elif ts_key not in seen_general and not no_kings:
                        seen_general.add(ts_key)
                        general_rows.append(row)
                        stats["passed"] += 1
                    if no_kings and etype_raw not in EXCLUDE_TYPES:
                        # Deduplicate by location + timeslot + property.
                        # Multiple organizers often post separate Mobilize events
                        # for the same physical protest; collapse them to one row.
                        lat, lon = row["event_lat"], row["event_lon"]
                        if lat != "" and lon != "":
                            loc_key = (round(float(lat), 4), round(float(lon), 4))
                        else:
                            loc_key = row["event_location"]  # fallback for virtual
                        nk_key = (loc_key, row["event_dt_sort"], prop["id"])
                        if nk_key not in seen_no_kings:
                            seen_no_kings.add(nk_key)
                            no_kings_rows.append(row)

    # ── First pass ────────────────────────────────────────────────────────────
    retry_queue: list[list[dict]] = []   # clusters that hit 429 on first pass

    bar = (_tqdm(clusters, unit="cluster", dynamic_ncols=True, colour="cyan")
           if _HAS_TQDM else None)
    cluster_iter = bar if bar is not None else clusters

    for idx, cluster in enumerate(cluster_iter, 1):
        zip_code, query_radius = cluster_query_params(cluster)
        names = ", ".join(p["name"] for p in cluster[:3])
        suffix = f" +{len(cluster) - 3} more" if len(cluster) > 3 else ""
        label = f"zip={zip_code} r={query_radius}mi  {names}{suffix}"

        if bar is not None:
            bar.set_description(label)
        else:
            print(f"  [{idx:>3}/{len(clusters)}] {label}")

        try:
            events = fetch_events_for_zip(zip_code, query_radius, now_ts, end_30d)
            if bar is not None:
                bar.set_postfix(events=len(events), refresh=True)
            else:
                print(f"           → {len(events)} event(s) returned")
            process_events(events, cluster)
        except RateLimitError:
            print(f"\n    [429] Rate-limited — queued for retry pass.", file=sys.stderr)
            retry_queue.append(cluster)

    if bar is not None:
        bar.close()

    # ── Retry pass (60 s head-start, then 5 s between attempts) ─────────────
    if retry_queue:
        print(f"\n  {len(retry_queue)} cluster(s) rate-limited. "
              f"Waiting 60s before retry pass …")
        time.sleep(60)

        still_failed: list[str] = []
        for attempt_num in range(1, MAX_RETRIES + 1):
            next_retry: list[list[dict]] = []

            for cluster in retry_queue:
                zip_code, query_radius = cluster_query_params(cluster)
                names = cluster[0]["name"]
                print(f"  [retry {attempt_num}/{MAX_RETRIES}] zip={zip_code}  ({names})")
                try:
                    events = fetch_events_for_zip(zip_code, query_radius, now_ts, end_30d)
                    print(f"           → {len(events)} event(s) returned")
                    process_events(events, cluster)
                    time.sleep(5)
                except RateLimitError:
                    next_retry.append(cluster)

            retry_queue = next_retry
            if not retry_queue:
                print("  All retries succeeded.")
                break

            if attempt_num < MAX_RETRIES:
                wait = 60 * attempt_num
                print(f"  {len(retry_queue)} still rate-limited. "
                      f"Waiting {wait}s before next retry …")
                time.sleep(wait)

        if retry_queue:
            for cluster in retry_queue:
                zip_code, _ = cluster_query_params(cluster)
                still_failed.append(zip_code)
                for prop in cluster:
                    print(f"  [FAILED] {prop['name']} (zip={prop['zip']}) — "
                          f"no data after {MAX_RETRIES} retries.", file=sys.stderr)

    total_checked = sum(stats.values())
    print(f"\n  ── Filter breakdown (event × property pairs) ──────────────────")
    print(f"     No coordinates (virtual/TBD) : {stats['no_coords']:>6}")
    print(f"     Outside 3-mile radius        : {stats['too_far']:>6}")
    print(f"     Excluded event type          : {stats['excluded_type']:>6}")
    print(f"     Outside 3-day window         : {stats['outside_window']:>6}")
    print(f"     ✓ Passed to 3-day sheet      : {stats['passed']:>6}")
    print(f"  ───────────────────────────────────────────────────────────────\n")

    general_rows.sort(key=lambda r: r["event_dt_sort"])
    no_kings_rows.sort(key=lambda r: r["event_dt_sort"])
    return general_rows, no_kings_rows


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _cell(ws, row: int, col: int, value,
          font=None, fill=None, align=None, border=None, hyperlink=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:    c.font      = font
    if fill:    c.fill      = fill
    if align:   c.alignment = align
    if border:  c.border    = border
    if hyperlink:
        c.hyperlink = hyperlink
    return c


def write_data_sheet(ws, rows: list[dict],
                     title_fill: PatternFill,
                     col_hdr_fill: PatternFill,
                     alt_fill: PatternFill,
                     sheet_title: str,
                     subtitle: str) -> None:

    ws.sheet_view.showGridLines = False
    num_cols = len(COLUMNS)
    last_col = get_column_letter(num_cols)
    notes_col_idx = next((i + 1 for i, (_, k, _) in enumerate(COLUMNS)
                          if k == "notes"), None)

    # Sort closest → furthest
    rows = sorted(rows, key=lambda r: r.get("distance_mi", 9999))

    def _dist_band(d):
        return "red" if d < 1.0 else ("amber" if d < 2.0 else "green")

    # ── Row 1: Title ──
    ws.merge_cells(f"A1:{last_col}1")
    _cell(ws, 1, 1, sheet_title,
          font   = Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
          fill   = title_fill,
          align  = Alignment(horizontal="center", vertical="center"),
          border = THIN)
    ws.row_dimensions[1].height = 26

    # ── Row 2: Subtitle ──
    ws.merge_cells(f"A2:{last_col}2")
    _cell(ws, 2, 1, subtitle,
          font   = Font(name="Calibri", italic=True, size=9, color="FFFFFF"),
          fill   = title_fill,
          align  = Alignment(horizontal="center", vertical="center"),
          border = THIN)
    ws.row_dimensions[2].height = 14

    # ── Row 3: Column headers + auto-filter ──
    for col_idx, (label, key, _) in enumerate(COLUMNS, 1):
        is_notes_hdr = (key == "notes")
        _cell(ws, 3, col_idx, label,
              font   = Font(name="Calibri", bold=True, size=9,
                            color="FFFFFF" if not is_notes_hdr else "FFFFFF",
                            italic=is_notes_hdr),
              fill   = col_hdr_fill,
              align  = Alignment(horizontal="center", vertical="center", wrap_text=True),
              border = THIN)
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}{max(4, 3 + len(rows))}"

    # ── Data rows ──
    if not rows:
        ws.merge_cells(f"A4:{last_col}4")
        _cell(ws, 4, 1, "No events found for this reporting period.",
              font  = Font(name="Calibri", italic=True, size=11, color="808080"),
              align = Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[4].height = 24
    else:
        col_widths = [len(label) for label, _, _ in COLUMNS]

        for i, row in enumerate(rows):
            row_idx = i + 4
            dist = row.get("distance_mi", 9999)
            band = _dist_band(dist)

            # Group edge detection
            prev_band = _dist_band(rows[i - 1].get("distance_mi", 9999)) if i > 0 else None
            next_band = _dist_band(rows[i + 1].get("distance_mi", 9999)) if i < len(rows) - 1 else None
            top_edge = (band != prev_band)
            bot_edge = (band != next_band)

            row_fill = {"red": FILL_RED, "amber": FILL_AMBER, "green": FILL_GREEN}[band]
            ws.row_dimensions[row_idx].height = 14

            for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
                val      = row.get(key, "")
                is_url   = (key == "event_url")
                is_notes = (key == "notes")
                is_dist  = (key == "distance_mi")
                is_prop  = (key == "property_name")
                is_left  = (col_idx == 1)
                is_right = (col_idx == num_cols)

                cell_fill = FILL_NOTES if is_notes else row_fill

                if is_url and val:
                    display = "Link"
                    font = Font(name="Calibri", size=9, bold=False,
                                color="1155CC", underline="single")
                elif is_prop:
                    display = val
                    font = Font(name="Calibri", size=9, bold=True)
                elif is_notes:
                    display = val
                    font = Font(name="Calibri", size=9, italic=True, color="888888")
                else:
                    display = val
                    font = Font(name="Calibri", size=9)

                align = Alignment(
                    vertical   = "center",
                    horizontal = "center" if is_dist else "left",
                    wrap_text  = False,
                )
                border = _group_border(top_edge, bot_edge, is_left, is_right)
                c = _cell(ws, row_idx, col_idx, display,
                          font=font, fill=cell_fill, align=align, border=border)
                if is_url and val:
                    c.hyperlink = val

                if not is_notes and display:
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1],
                                                   len(str(display)))

        # ── Auto-size columns ──
        for col_idx, (_, _, max_width) in enumerate(COLUMNS, 1):
            fitted = min(col_widths[col_idx - 1] + 2, max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(fitted, 8)

        # ── Color legend ──
        legend_row = 4 + len(rows) + 2
        distance_items = [
            ("  Within 1 mile",    FILL_RED),
            ("  1 – 2 miles",      FILL_AMBER),
            ("  Beyond 2 miles",   FILL_GREEN),
            ("  Analyst notes",    FILL_NOTES),
        ]
        _cell(ws, legend_row, 1, "Distance Key:",
              font  = Font(name="Calibri", bold=True, size=9, color="2C3E50"),
              align = Alignment(vertical="center"))
        ws.row_dimensions[legend_row].height = 14
        for j, (label, fill) in enumerate(distance_items):
            r = legend_row + 1 + j
            swatch = ws.cell(row=r, column=1)
            swatch.fill = fill
            swatch.border = _group_border(True, True, True, True)
            ws.row_dimensions[r].height = 14
            _cell(ws, r, 2, label,
                  font  = Font(name="Calibri", size=9),
                  align = Alignment(vertical="center"))

        # ── Flag legend ──
        flag_row = legend_row + len(distance_items) + 2
        _cell(ws, flag_row, 1, "Flag Columns:",
              font  = Font(name="Calibri", bold=True, size=9, color="2C3E50"),
              align = Alignment(vertical="center"))
        ws.row_dimensions[flag_row].height = 14
        flag_items = [
            ("Is New?",      "Event ID not seen in the previous run's data"),
            ("Duplicate?",   "Same event appears near 2+ tracked properties this run"),
            ("Recurring?",   "Same event has 2+ distinct time slots this run"),
        ]
        for j, (col_name, description) in enumerate(flag_items):
            r = flag_row + 1 + j
            ws.row_dimensions[r].height = 14
            _cell(ws, r, 1, col_name,
                  font  = Font(name="Calibri", bold=True, size=9),
                  align = Alignment(vertical="center"))
            _cell(ws, r, 2, description,
                  font  = Font(name="Calibri", size=9),
                  align = Alignment(vertical="center"))


def build_summary_sheet(ws, general_rows: list[dict],
                         no_kings_rows: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    now_str = datetime.now(_EASTERN).strftime("%B %d, %Y at %H:%M EST")

    # Title
    ws.merge_cells("A1:C1")
    _cell(ws, 1, 1, "PROTEST TRACKER — EXECUTIVE SUMMARY",
          font   = Font(name="Calibri", bold=True, size=18, color="FFFFFF"),
          fill   = HDR_NAVY,
          align  = Alignment(horizontal="center", vertical="center"),
          border = THIN)
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:C2")
    _cell(ws, 2, 1, f"Report generated: {now_str}",
          font   = Font(name="Calibri", italic=True, size=11, color="FFFFFF"),
          fill   = HDR_NAVY,
          align  = Alignment(horizontal="center", vertical="center"),
          border = THIN)
    ws.row_dimensions[2].height = 20

    # Section header
    _cell(ws, 3, 1, "Metric",
          font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
          fill  = HDR_BLUE,
          align = Alignment(horizontal="center", vertical="center"),
          border = THIN)
    _cell(ws, 3, 2, "Value",
          font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
          fill  = HDR_BLUE,
          align = Alignment(horizontal="center", vertical="center"),
          border = THIN)
    ws.merge_cells("B3:C3")
    ws.row_dimensions[3].height = 20

    # Unique property IDs with hits
    gen_props  = {r["property_id"] for r in general_rows}
    nk_props   = {r["property_id"] for r in no_kings_rows}
    gen_events = {(r["event_title"], r["event_date"]) for r in general_rows}
    nk_events  = {(r["event_title"], r["event_date"]) for r in no_kings_rows}

    end_3d  = (datetime.now() + timedelta(days=DAYS_GENERAL)).strftime("%b %d, %Y")
    end_30d = (datetime.now() + timedelta(days=DAYS_NO_KINGS)).strftime("%b %d, %Y")

    stats = [
        ("Properties monitored",                  "205"),
        ("Search radius",                          f"{SEARCH_RADIUS_MI} miles"),
        ("─── 3-Day Window ───",                  ""),
        (f"  Window end date",                     end_3d),
        (f"  Properties with events",              str(len(gen_props))),
        (f"  Unique events found",                 str(len(gen_events))),
        (f"  Total property-event pairs",          str(len(general_rows))),
        ("─── No Kings 30-Day Window ───",        ""),
        (f"  Window end date",                     end_30d),
        (f"  Properties with No Kings events",     str(len(nk_props))),
        (f"  Unique No Kings events found",        str(len(nk_events))),
        (f"  Total property-event pairs",          str(len(no_kings_rows))),
        ("─── Data Sources ───",                  ""),
        ("  Primary source",                       "Mobilize.us Public API"),
        ("  Keywords (No Kings filter)",           ", ".join(NO_KINGS_KEYWORDS)),
    ]

    for i, (metric, value) in enumerate(stats, start=4):
        is_even   = (i % 2 == 0)
        is_hdr    = metric.startswith("───")
        row_fill  = (PatternFill("solid", fgColor="D9E2F3") if is_hdr
                     else (ALT_BLUE if is_even else WHITE_FILL))
        bold      = is_hdr

        _cell(ws, i, 1, metric,
              font  = Font(name="Calibri", bold=bold, size=11,
                           color="1F3864" if is_hdr else "000000"),
              fill  = row_fill,
              align = Alignment(vertical="center"),
              border = THIN)
        ws.merge_cells(f"B{i}:C{i}")
        _cell(ws, i, 2, value,
              font  = Font(name="Calibri", bold=bold, size=11),
              fill  = row_fill,
              align = Alignment(vertical="center"),
              border = THIN)
        ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 5


# ── Cache helpers ──────────────────────────────────────────────────────────────

CACHE_FILE = "protest_tracker_cache.json"


def save_cache(general_rows: list[dict], no_kings_rows: list[dict],
               path: str = CACHE_FILE) -> None:
    def _serialize(rows):
        out = []
        for r in rows:
            d = dict(r)
            if isinstance(d.get("event_dt_sort"), datetime):
                d["event_dt_sort"] = d["event_dt_sort"].isoformat()
            out.append(d)
        return out

    with open(path, "w") as f:
        json.dump({"general": _serialize(general_rows),
                   "no_kings": _serialize(no_kings_rows)}, f)
    print(f"  ✓ Data cache saved → {path}")


def load_cache(path: str = CACHE_FILE) -> tuple[list[dict], list[dict]]:
    with open(path) as f:
        data = json.load(f)

    def _deserialize(rows):
        out = []
        for r in rows:
            d = dict(r)
            if isinstance(d.get("event_dt_sort"), str):
                d["event_dt_sort"] = datetime.fromisoformat(d["event_dt_sort"])
            out.append(d)
        return out

    return _deserialize(data["general"]), _deserialize(data["no_kings"])


# ── Build workbook ─────────────────────────────────────────────────────────────

def build_excel(general_rows: list[dict], no_kings_rows: list[dict],
                output_path: str) -> None:
    wb = Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    build_summary_sheet(ws_sum, general_rows, no_kings_rows)

    # 3-day events sheet
    ws_gen = wb.create_sheet("3-Day Events")
    now_str = datetime.now().strftime("%B %d, %Y")
    end_3d  = (datetime.now() + timedelta(days=DAYS_GENERAL)).strftime("%B %d, %Y")
    write_data_sheet(
        ws_gen, general_rows,
        title_fill   = HDR_NAVY,
        col_hdr_fill = HDR_BLUE,
        alt_fill     = ALT_BLUE,
        sheet_title  = "Protest / Rally Events Near Tracked Properties — 3-Day Window",
        subtitle     = (f"{now_str}  through  {end_3d}"
                        f"  •  Within {SEARCH_RADIUS_MI} miles of property"
                        f"  •  Source: Mobilize.us"),
    )

    # No Kings sheet
    ws_nk = wb.create_sheet("No Kings Events (30-Day)")
    end_30d = (datetime.now() + timedelta(days=DAYS_NO_KINGS)).strftime("%B %d, %Y")
    write_data_sheet(
        ws_nk, no_kings_rows,
        title_fill   = HDR_PURPLE,
        col_hdr_fill = COL_PURPLE,
        alt_fill     = ALT_PURPLE,
        sheet_title  = "No Kings Protest Events — 30-Day Window",
        subtitle     = (f"{now_str}  through  {end_30d}"
                        f"  •  Within {SEARCH_RADIUS_MI} miles of property"
                        f"  •  Source: Mobilize.us"),
    )

    wb.save(output_path)
    print(f"\n  ✓ Excel workbook saved → {output_path}")


def update_trend_data(general_rows: list[dict], no_kings_rows: list[dict],
                       path: str = "trend_data.json") -> None:
    """Append a summary entry to trend_data.json for historical charting."""
    import json as _json

    def _band_counts(rows):
        red   = sum(1 for r in rows if (float(r.get("distance_mi") or 9999)) < 1)
        amber = sum(1 for r in rows if 1 <= (float(r.get("distance_mi") or 9999)) < 2)
        green = sum(1 for r in rows if (float(r.get("distance_mi") or 9999)) >= 2)
        new   = sum(1 for r in rows if str(r.get("is_new", "")).lower() == "yes")
        return red, amber, green, new

    now = datetime.now(_EASTERN)
    gr, ga, gg, gn = _band_counts(general_rows)
    nr, na, ng, nn = _band_counts(no_kings_rows)
    entry = {
        "ts":        now.isoformat(timespec="seconds"),
        "label":     now.strftime("%b %d"),
        "gen_total": len(general_rows),
        "gen_new":   gn,
        "gen_red":   gr, "gen_amber": ga, "gen_green": gg,
        "nk_total":  len(no_kings_rows),
        "nk_new":    nn,
        "nk_red":    nr, "nk_amber":  na, "nk_green":  ng,
    }
    try:
        with open(path) as f:
            data = _json.load(f)
    except (FileNotFoundError, ValueError):
        data = {"runs": []}
    data["runs"].append(entry)
    data["runs"] = data["runs"][-270:]          # keep last 270 runs (~3 months at 3×/day)
    with open(path, "w") as f:
        _json.dump(data, f)
    print(f"  ✓ Trend data updated      → {path}  ({len(data['runs'])} runs)")


def build_dashboard_json(general_rows: list[dict], no_kings_rows: list[dict],
                          output_path: str = "dashboard_data.json") -> None:
    import json as _json
    now = datetime.now(_EASTERN)
    end_3d  = (now + timedelta(days=DAYS_GENERAL)).strftime("%B %d, %Y")
    end_30d = (now + timedelta(days=DAYS_NO_KINGS)).strftime("%B %d, %Y")

    _DASHBOARD_EXTRA = ["event_lat", "event_lon", "prop_lat", "prop_lon", "sponsor_name"]

    def _clean(rows):
        out = []
        for r in sorted(rows, key=lambda x: x.get("distance_mi", 9999)):
            d = {k: r.get(k, "") for _, k, _ in COLUMNS}
            for key in _DASHBOARD_EXTRA:
                d[key] = r.get(key, "")
            out.append(d)
        return out

    payload = {
        "generated":     now.strftime("%B %d, %Y at %H:%M EST"),
        "generated_iso": now.astimezone(timezone.utc).isoformat(),
        "general": {
            "title": "3-Day Window",
            "subtitle": f"{now.strftime('%B %d, %Y')}  through  {end_3d}  •  Within {SEARCH_RADIUS_MI} miles",
            "rows": _clean(general_rows),
        },
        "no_kings": {
            "title": "No Kings — 30-Day Window",
            "subtitle": f"{now.strftime('%B %d, %Y')}  through  {end_30d}  •  Within {SEARCH_RADIUS_MI} miles",
            "rows": _clean(no_kings_rows),
        },
    }
    with open(output_path, "w") as f:
        _json.dump(payload, f)
    print(f"  ✓ Dashboard JSON saved  → {output_path}")
    update_trend_data(general_rows, no_kings_rows)


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Protest Tracker — generates an Excel report of events near tracked properties",
    )
    ap.add_argument(
        "--output", "-o",
        default="protest_tracker_report.xlsx",
        help="Output Excel file name (default: protest_tracker_report.xlsx)",
    )
    ap.add_argument(
        "--csv", "-c",
        default=PROPERTIES_CSV,
        help=f"Path to properties CSV (default: {PROPERTIES_CSV})",
    )
    ap.add_argument(
        "--api-key", "-k",
        default="",
        help="Mobilize.us API key (overrides MOBILIZE_API_KEY env var)",
    )
    ap.add_argument(
        "--regen", "-r",
        action="store_true",
        help=f"Skip API calls and rebuild Excel from cached data ({CACHE_FILE})",
    )
    ap.add_argument(
        "--cache", default=CACHE_FILE,
        help=f"Path to cache file (default: {CACHE_FILE})",
    )
    args = ap.parse_args()

    # Allow CLI flag to override env var
    if args.api_key:
        global MOBILIZE_API_KEY
        MOBILIZE_API_KEY = args.api_key

    print("=" * 64)
    print("  PROTEST TRACKER")
    print("=" * 64)

    if args.regen:
        print(f"  Mode           : REGEN (rebuilding from cache)")
        print(f"  Cache file     : {args.cache}")
        print(f"  Output file    : {args.output}")
        print("=" * 64)
        if not os.path.exists(args.cache):
            print(f"\n  ERROR: cache file not found: {args.cache}", file=sys.stderr)
            sys.exit(1)
        print(f"\nLoading cached data from {args.cache} …")
        general_rows, no_kings_rows = load_cache(args.cache)
        print(f"  3-Day events   : {len(general_rows)} rows")
        print(f"  No Kings events: {len(no_kings_rows)} rows")
        # Collapse in case this cache predates the timeslot-collapse feature
        general_rows  = collapse_recurring_timeslots(general_rows)
        no_kings_rows = collapse_recurring_timeslots(no_kings_rows)
    else:
        print(f"  Properties CSV : {args.csv}")
        print(f"  Output file    : {args.output}")
        print(f"  Authenticated  : {'Yes' if MOBILIZE_API_KEY else 'No (set MOBILIZE_API_KEY for better rate limits)'}")
        print(f"  Search radius  : {SEARCH_RADIUS_MI} miles")
        print(f"  Cluster radius : {CLUSTER_RADIUS_MI} miles (properties grouped for fewer API calls)")
        print(f"  General window : today + {DAYS_GENERAL} days")
        print(f"  No Kings window: today + {DAYS_NO_KINGS} days")
        print("=" * 64)

        properties = load_properties(args.csv)
        print(f"\nLoaded {len(properties)} properties.\n")
        print("Querying Mobilize.us public API …\n")

        # Load previous cache (if downloaded by the workflow) for is_new comparison
        prev_event_ids = get_prev_event_ids(args.cache)
        if prev_event_ids:
            print(f"  Previous cache : {len(prev_event_ids)} known event IDs loaded for new-event detection")

        general_rows, no_kings_rows = collect_events(properties)

        print(f"\n{'─'*64}")
        print(f"  3-Day events found        : {len(general_rows)} property-event rows")
        print(f"  No Kings events found     : {len(no_kings_rows)} property-event rows")
        print(f"{'─'*64}")

        print("\nAnnotating event flags (new / duplicate / recurring) …")
        annotate_event_flags(general_rows,  prev_event_ids)
        annotate_event_flags(no_kings_rows, prev_event_ids)
        new_g  = sum(1 for r in general_rows  if r.get("is_new"))
        new_nk = sum(1 for r in no_kings_rows if r.get("is_new"))
        dup_g  = sum(1 for r in general_rows  if r.get("is_duplicate"))
        rec_g  = sum(1 for r in general_rows  if r.get("is_recurring"))
        print(f"  New events (3-day)        : {new_g}")
        print(f"  New events (No Kings)     : {new_nk}")
        print(f"  Duplicates (3-day)        : {dup_g}")
        print(f"  Recurring  (3-day)        : {rec_g}")

        print("\nCollapsing recurring events to nearest timeslot …")
        general_rows  = collapse_recurring_timeslots(general_rows)
        no_kings_rows = collapse_recurring_timeslots(no_kings_rows)
        print(f"  3-Day rows after collapse : {len(general_rows)}")
        print(f"  No Kings rows after collapse: {len(no_kings_rows)}")

        save_cache(general_rows, no_kings_rows, args.cache)

    print("\nBuilding Excel workbook …")
    build_excel(general_rows, no_kings_rows, args.output)
    build_dashboard_json(general_rows, no_kings_rows)
    print("Done.\n")


if __name__ == "__main__":
    main()
